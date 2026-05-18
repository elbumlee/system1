import os
from pathlib import Path
from typing import List, Optional
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models import Game
from storage.base import AbstractStorage


HEADERS = ["ID", "Name", "Steam", "Epic", "Switch", "Added Date", "Notes", "Genre", "Favorite"]


class ExcelStorage(AbstractStorage):
    def __init__(self, file_path: Optional[str] = None):
        if file_path:
            self.file_path = Path(os.path.expanduser(file_path))
        else:
            self.file_path = Path.home() / "game_library.xlsx"
        self._ensure_file()

    def _ensure_file(self):
        """Create the Excel file with headers if it doesn't exist."""
        if not self.file_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Games"

            # Style the header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, header in enumerate(HEADERS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Set column widths
            col_widths = [36, 40, 8, 8, 8, 15, 40, 20, 10]
            for col_idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

            wb.save(self.file_path)

    def _load_workbook(self):
        return load_workbook(self.file_path)

    def get_all_games(self) -> List[Game]:
        """Read all games from the Excel file."""
        wb = self._load_workbook()
        ws = wb.active
        games = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            # Support old files with 7 columns (no Genre/Favorite) by checking row length
            game_id = row[0]
            name = row[1] if len(row) > 1 else None
            steam = row[2] if len(row) > 2 else None
            epic = row[3] if len(row) > 3 else None
            switch = row[4] if len(row) > 4 else None
            added_date = row[5] if len(row) > 5 else None
            notes = row[6] if len(row) > 6 else None
            genre = row[7] if len(row) > 7 else None
            favorite = row[8] if len(row) > 8 else None

            games.append(Game(
                id=str(game_id),
                name=str(name) if name else "",
                steam=bool(steam) if steam is not None else False,
                epic=bool(epic) if epic is not None else False,
                switch=bool(switch) if switch is not None else False,
                added_date=str(added_date) if added_date else date.today().isoformat(),
                notes=str(notes) if notes else "",
                genre=str(genre) if genre else "",
                favorite=bool(favorite) if favorite is not None else False,
            ))

        return games

    def add_game(self, game: Game) -> Game:
        """Append a new game row to the Excel file."""
        wb = self._load_workbook()
        ws = wb.active

        # Find next empty row
        next_row = ws.max_row + 1

        ws.cell(row=next_row, column=1, value=game.id)
        ws.cell(row=next_row, column=2, value=game.name)
        ws.cell(row=next_row, column=3, value=game.steam)
        ws.cell(row=next_row, column=4, value=game.epic)
        ws.cell(row=next_row, column=5, value=game.switch)
        ws.cell(row=next_row, column=6, value=game.added_date)
        ws.cell(row=next_row, column=7, value=game.notes)
        ws.cell(row=next_row, column=8, value=game.genre)
        ws.cell(row=next_row, column=9, value=game.favorite)

        wb.save(self.file_path)
        return game

    def update_game(self, game_id: str, updates: dict) -> Optional[Game]:
        """Update a game's fields in the Excel file."""
        wb = self._load_workbook()
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value == game_id:
                if "name" in updates and updates["name"] is not None:
                    ws.cell(row=row_idx, column=2, value=updates["name"])
                if "steam" in updates and updates["steam"] is not None:
                    ws.cell(row=row_idx, column=3, value=updates["steam"])
                if "epic" in updates and updates["epic"] is not None:
                    ws.cell(row=row_idx, column=4, value=updates["epic"])
                if "switch" in updates and updates["switch"] is not None:
                    ws.cell(row=row_idx, column=5, value=updates["switch"])
                if "notes" in updates and updates["notes"] is not None:
                    ws.cell(row=row_idx, column=7, value=updates["notes"])
                if "genre" in updates and updates["genre"] is not None:
                    ws.cell(row=row_idx, column=8, value=updates["genre"])
                if "favorite" in updates and updates["favorite"] is not None:
                    ws.cell(row=row_idx, column=9, value=updates["favorite"])
                wb.save(self.file_path)

                # Reload and return updated game
                return self.get_game_by_id(game_id)

        return None

    def delete_game(self, game_id: str) -> bool:
        """Delete a game row from the Excel file."""
        wb = self._load_workbook()
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value == game_id:
                ws.delete_rows(row_idx)
                wb.save(self.file_path)
                return True

        return False

    def get_file_path(self) -> str:
        return str(self.file_path)
