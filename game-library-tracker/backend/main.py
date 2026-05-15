import io
import os
import uuid
from datetime import date
from typing import Dict, List, Optional

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse

load_dotenv()

from models import Game, GameCreate, GameUpdate, OCRResult, OCRConfirm, StorageConfig, StorageType
from storage.excel_storage import ExcelStorage
from storage.sheets_storage import SheetsStorage

# ---------------------------------------------------------------------------
# App initialisation
# ---------------------------------------------------------------------------

app = FastAPI(title="Game Library Tracker API", version="1.0.0")

# CORS – allow localhost on common dev ports AND any local-network address
origins = [
    "http://localhost",
    "http://localhost:3000",
    "http://localhost:5173",
    "http://localhost:4173",
    "http://127.0.0.1:5173",
    "http://127.0.0.1:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_origin_regex=r"http://192\.168\.\d+\.\d+(:\d+)?",  # local network
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Storage management
# ---------------------------------------------------------------------------

_storage_type: str = os.environ.get("STORAGE_TYPE", "excel")
_excel_file_path: str = os.environ.get("EXCEL_FILE_PATH", "~/game_library.xlsx")
_google_sheet_id: str = os.environ.get("GOOGLE_SHEET_ID", "")
_google_credentials_path: str = os.environ.get("GOOGLE_CREDENTIALS_PATH", "")

_storage = None  # lazy-init in get_storage()


def _create_storage():
    global _storage_type, _excel_file_path, _google_sheet_id, _google_credentials_path
    if _storage_type == "sheets":
        if not _google_sheet_id:
            raise HTTPException(status_code=400, detail="Google Sheet ID is not configured.")
        return SheetsStorage(
            sheet_id=_google_sheet_id,
            credentials_path=_google_credentials_path or None,
        )
    else:
        return ExcelStorage(file_path=_excel_file_path or None)


def get_storage():
    global _storage
    if _storage is None:
        _storage = _create_storage()
    return _storage


# ---------------------------------------------------------------------------
# In-memory OCR result cache (image_id -> OCRResult)
# ---------------------------------------------------------------------------

_ocr_cache: Dict[str, OCRResult] = {}

# ---------------------------------------------------------------------------
# Routes: Games
# ---------------------------------------------------------------------------


@app.get("/api/games", response_model=List[Game])
def list_games():
    """Return all games from the current storage backend."""
    storage = get_storage()
    return storage.get_all_games()


@app.post("/api/games", response_model=Game, status_code=201)
def create_game(payload: GameCreate):
    """Add a new game to the library."""
    storage = get_storage()
    game = Game(
        id=str(uuid.uuid4()),
        name=payload.name,
        steam=payload.steam,
        epic=payload.epic,
        switch=payload.switch,
        added_date=date.today().isoformat(),
        notes=payload.notes,
    )
    return storage.add_game(game)


@app.put("/api/games/{game_id}", response_model=Game)
def update_game(game_id: str, payload: GameUpdate):
    """Update a game (e.g. toggle platform checkbox)."""
    storage = get_storage()
    updated = storage.update_game(game_id, payload.model_dump(exclude_none=True))
    if updated is None:
        raise HTTPException(status_code=404, detail=f"Game '{game_id}' not found.")
    return updated


@app.delete("/api/games/{game_id}", status_code=204)
def delete_game(game_id: str):
    """Delete a game from the library."""
    storage = get_storage()
    deleted = storage.delete_game(game_id)
    if not deleted:
        raise HTTPException(status_code=404, detail=f"Game '{game_id}' not found.")


# ---------------------------------------------------------------------------
# Routes: Storage
# ---------------------------------------------------------------------------


@app.get("/api/storage/type", response_model=StorageType)
def get_storage_type():
    """Return the current storage backend and its configuration."""
    return StorageType(
        storage_type=_storage_type,
        excel_file_path=_excel_file_path if _storage_type == "excel" else None,
        google_sheet_id=_google_sheet_id if _storage_type == "sheets" else None,
    )


@app.post("/api/storage/switch")
def switch_storage(config: StorageConfig):
    """Switch the storage backend (excel <-> sheets)."""
    global _storage, _storage_type, _excel_file_path, _google_sheet_id, _google_credentials_path

    _storage_type = config.storage_type
    if config.excel_file_path:
        _excel_file_path = config.excel_file_path
    if config.google_sheet_id:
        _google_sheet_id = config.google_sheet_id
    if config.google_credentials_path:
        _google_credentials_path = config.google_credentials_path

    # Force re-init on next request
    _storage = None

    try:
        # Test the new connection eagerly
        new_storage = get_storage()
        new_storage.get_all_games()
        return {"status": "ok", "storage_type": _storage_type}
    except Exception as exc:
        _storage = None
        raise HTTPException(status_code=400, detail=str(exc))


# ---------------------------------------------------------------------------
# Routes: OCR
# ---------------------------------------------------------------------------


@app.post("/api/ocr/upload", response_model=OCRResult)
async def upload_image_for_ocr(file: UploadFile = File(...)):
    """Accept an image, run OCR, return candidate game names."""
    try:
        from ocr.processor import OCRProcessor
        processor = OCRProcessor()
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc))

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Uploaded file must be an image.")

    image_bytes = await file.read()
    if len(image_bytes) == 0:
        raise HTTPException(status_code=400, detail="Uploaded image is empty.")

    try:
        candidates, platform_hint = processor.process(image_bytes)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {exc}")

    image_id = str(uuid.uuid4())
    result = OCRResult(image_id=image_id, candidates=candidates, platform_hint=platform_hint)
    _ocr_cache[image_id] = result
    return result


@app.post("/api/ocr/confirm", response_model=List[Game])
def confirm_ocr_games(payload: OCRConfirm):
    """Confirm selected game names from an OCR result and add them to the library."""
    if payload.image_id not in _ocr_cache:
        raise HTTPException(status_code=404, detail="OCR result not found. Please re-upload the image.")

    storage = get_storage()
    added_games: List[Game] = []
    today = date.today().isoformat()

    for name in payload.selected_names:
        name = name.strip()
        if not name:
            continue

        steam = payload.platform == "steam"
        epic = payload.platform == "epic"
        switch = payload.platform == "switch"

        game = Game(
            id=str(uuid.uuid4()),
            name=name,
            steam=steam,
            epic=epic,
            switch=switch,
            added_date=today,
            notes="Added via OCR",
        )
        storage.add_game(game)
        added_games.append(game)

    # Remove from cache once confirmed
    del _ocr_cache[payload.image_id]
    return added_games


# ---------------------------------------------------------------------------
# Routes: Export
# ---------------------------------------------------------------------------


@app.get("/api/export/excel")
def export_excel():
    """Download the current game library as an Excel file."""
    import tempfile
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    games = get_storage().get_all_games()

    wb = Workbook()
    ws = wb.active
    ws.title = "Games"

    headers = ["ID", "Name", "Steam", "Epic", "Switch", "Added Date", "Notes"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    col_widths = [36, 40, 8, 8, 8, 15, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    for row_idx, game in enumerate(games, start=2):
        ws.cell(row=row_idx, column=1, value=game.id)
        ws.cell(row=row_idx, column=2, value=game.name)
        ws.cell(row=row_idx, column=3, value=game.steam)
        ws.cell(row=row_idx, column=4, value=game.epic)
        ws.cell(row=row_idx, column=5, value=game.switch)
        ws.cell(row=row_idx, column=6, value=game.added_date)
        ws.cell(row=row_idx, column=7, value=game.notes)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=game_library.xlsx"},
    )


# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------


@app.get("/api/health")
def health():
    return {"status": "ok", "storage_type": _storage_type}


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
