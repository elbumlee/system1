import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models import Game


def generate_excel(games: List[Game]) -> io.BytesIO:
    """Generate an Excel workbook from a list of games and return it as a BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Games"

    headers = ["ID", "Name", "Steam", "Epic", "Switch", "Added Date", "Notes"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    col_widths = [36, 40, 8, 8, 8, 15, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    for row_idx, game in enumerate(games, start=2):
        ws.cell(row=row_idx, column=1, value=game.id)
        ws.cell(row=row_idx, column=2, value=game.name)
        ws.cell(row=row_idx, column=3, value=game.steam)
        ws.cell(row=row_idx, column=4, value=game.epic)
        ws.cell(row=row_idx, column=5, value=game.switch)
        ws.cell(row=row_idx, column=6, value=game.added_date)
        ws.cell(row=row_idx, column=7, value=game.notes)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
